import requests
import os
import pandas as pd
import time
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

USERNAME = ""
PASSWORD = ""

CAS_BASE_URL = "https://cas.epias.com.tr/cas/v1"
SERVICE_URL = "https://epys.epias.com.tr/pre-reconciliation/v1/meter-data/approved-meter-data/hourly/export"
SERVICE_NAME = SERVICE_URL

OUTPUT_DIR = "saatlik"
MAIN_FILE = "main.xlsx"
MAX_WORKERS = 20  # Aynı anda kaç sayaç indirileceği


# ------------------------------------------------------------
# Retry destekli session oluştur
# ------------------------------------------------------------
def create_retry_session():
    session = requests.Session()

    retry = Retry(
        total=5,
        backoff_factor=1,  # 1 → 2 → 4 → 8 saniye
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["POST"],
    )

    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    return session


# ------------------------------------------------------------
# CAS TGT alma
# ------------------------------------------------------------
def get_tgt():
    url = f"{CAS_BASE_URL}/tickets"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {"username": USERNAME, "password": PASSWORD}

    response = requests.post(url, headers=headers, data=data, allow_redirects=False)
    response.raise_for_status()

    tgt_location = response.headers.get("Location", "")
    if "TGT" not in tgt_location:
        raise Exception(f"TGT alınamadı! Dönen header: {tgt_location}")

    return tgt_location.split("/")[-1].strip()


# ------------------------------------------------------------
# CAS ST alma
# ------------------------------------------------------------
def get_st(session, tgt):
    url = f"{CAS_BASE_URL}/tickets/{tgt}"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {"service": SERVICE_NAME}
    response = session.post(url, headers=headers, data=data)
    response.raise_for_status()
    return response.text.strip()


# ------------------------------------------------------------
# Sayaç indir (Retry destekli)
# ------------------------------------------------------------
def export_meter_data(session, tgt, st, payload, meter_id):

    url = f"{SERVICE_URL}?ticket={st}"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "TGT": tgt,
    }

    retry_count = 3

    for attempt in range(1, retry_count + 1):
        try:
            response = session.post(url, headers=headers, json=payload, timeout=60)

            if response.status_code == 200:
                content_type = response.headers.get("Content-Type", "")

                if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in content_type:
                    os.makedirs(OUTPUT_DIR, exist_ok=True)
                    filename = os.path.join(OUTPUT_DIR, f"onayli_sayac_{meter_id}.xlsx")

                    with open(filename, "wb") as f:
                        f.write(response.content)

                    print(f"Sayaç {meter_id} indirildi → {filename}")
                    return filename

            print(f"Sayaç {meter_id}: XLSX İçeriği gelmedi ({response.status_code})")

        except Exception as e:
            print(f"Sayaç {meter_id} hata ({attempt}/{retry_count}): {e}")
            time.sleep(3)

    print(f"Sayaç {meter_id} tamamen BAŞARISIZ")
    return None


# ------------------------------------------------------------
# HourlyPivot oluştur
# ------------------------------------------------------------
def create_hourly_pivot(main_file_path, file_list):
    wb = load_workbook(main_file_path)

    if "HourlyPivot" in wb.sheetnames:
        del wb["HourlyPivot"]

    new_sheet = wb.create_sheet("HourlyPivot")
    new_sheet.cell(row=1, column=1, value="Saat")

    column = 2
    for filename in file_list:
        if filename and os.path.exists(filename):

            hourly_wb = load_workbook(filename)
            sheet = hourly_wb.active

            meter_header = sheet["C2"].value or os.path.basename(filename)
            new_sheet.cell(row=1, column=column, value=meter_header)

            for row in range(2, sheet.max_row + 1):
                val = sheet[f"G{row}"].value
                try:
                    val = float(val) * 1000  # MWh -> kWh
                except:
                    val = 0
                new_sheet.cell(row=row, column=column, value=val)

            hourly_wb.close()
            column += 1

    wb.save(main_file_path)
    print("'HourlyPivot' sayfası oluşturuldu.")


# ------------------------------------------------------------
# MAIN
# ------------------------------------------------------------
if __name__ == "__main__":
    try:
        print("TGT alınıyor...")
        tgt = get_tgt()
        print(f"TGT: {tgt}")

        print("\nmain.xlsx okunuyor...")

        if not os.path.exists(MAIN_FILE):
            raise Exception(f"{MAIN_FILE} bulunamadı!")

        df = pd.read_excel(MAIN_FILE)

        if "Meter ID" in df.columns:
            id_col = "Meter ID"
        elif "SayacID" in df.columns:
            id_col = "SayacID"
        else:
            raise Exception("Excel'de 'Meter ID' veya 'SayacID' bulunamadı!")

        meter_ids = df[id_col].dropna().astype(int).tolist()
        print(f"{len(meter_ids)} sayaç bulundu, indiriliyor... (Thread={MAX_WORKERS})\n")

        downloaded_files = []
        session = create_retry_session()

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {}

            for mid in meter_ids:
                st = get_st(session, tgt)

                payload = {
                    "period": "2025-12-01T00:00:00+03:00",
                    "version": "2025-12-01T00:00:00+03:00",
                    "isRetrospective": False,
                    "region": "TR1",
                    "organization": 6464,
                    "isLastVersion": False,
                    "readStatus": True,
                    "exportType": "XLSX",
                    "meterId": mid,
                }

                futures[executor.submit(export_meter_data, session, tgt, st, payload, mid)] = mid

            for future in as_completed(futures):
                mid = futures[future]
                result = future.result()
                downloaded_files.append({"Meter ID": mid, "File Path": result or "File Not Exists."})
                print(f"Sayaç {mid} tamamlandı.")

        # EXCELE KAYDET
        pd.DataFrame(downloaded_files).to_excel(MAIN_FILE, sheet_name="ControlFile", index=False)
        print("'ControlFile' oluşturuldu.")

        # Pivot için geçerli dosyaları gönder
        valid_files = [d["File Path"] for d in downloaded_files if d["File Path"] and d["File Path"] != "File Not Exists."]
        create_hourly_pivot(MAIN_FILE, valid_files)

        print("\nTÜM İŞLEMLER BAŞARIYLA TAMAMLANDI ✅")
        print(f"Ana dosya: {os.path.abspath(MAIN_FILE)}")

    except Exception as e:
        print("\n HATA OLUŞTU:", e)
